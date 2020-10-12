from pyomo.environ import *
import os
import math
from openpyxl import *

wb=load_workbook(filename='14A.V-Mobile_CLASS.xlsx')
ws=wb["Data"]

model=AbstractModel()

#Size
model.ni=Param(initialize=3) #Number of carriers
model.nj=Param(initialize=5) #Destinations
model.nk=Param(initialize=3) #Price intervals
model.nt=Param(initialize=2) #Months

#Sets
model.i=RangeSet(1, model.ni) #Carriers
model.j=RangeSet(1, model.nj) #Destinations
model.k=RangeSet(1, model.nk) #Price intervals
model.t=RangeSet(1, model.nt) #Months

#Data
#Price
def p_init(model, i,j,k,t):
    return (ws.cell(17+(3*(i-1))+k, 1+(5*(t-1))+j).value)
model.p=Param(model.i, model.j, model.k ,model.t, initialize=p_init)

#Penalty
def pen_init(model, i,j,t):
    return (ws.cell(30+i, 1+(5*(t-1)+j)).value)
model.pen=Param(model.i, model.j, model.t, initialize=pen_init)

#Demand Forcast
def d_init(model,j,t):
    return (ws.cell(35, 5*(t-1)+j+1).value)
model.d=Param(model.j, model.t, initialize=d_init)

#Lower Threshold
def LT_init(model,i,k):
    return (ws.cell(5+(3*(i-1))+k, 2).value)
model.LT=Param(model.i, model.k, initialize=LT_init)

#Upper Threshold
def UT_init(model,i,k):
    return (ws.cell(5+(3*(i-1))+k, 3).value)
model.UT=Param(model.i, model.k, initialize=UT_init)

#Lower Bound
def LB_init(model,i,t):
    return (ws.cell(30+i, 14+(2*(t-1))).value)
model.LB=Param(model.i, model.t, initialize=LB_init)
            
#Upper Bound
def UB_init(model,i,t):
    return (ws.cell(30+i, 15+(2*(t-1))).value)
model.UB=Param(model.i, model.t, initialize=UB_init)

            
#Variables
model.X=Var(model.i, model.j, model.k, model.t, within=NonNegativeReals)
model.bin=Var(model.i, model.k, domain=Boolean)
model.Z=Var(model.i, model.j, model.k, model.t, within=NonNegativeReals)
            
            
#Expressions

            
#OBJECTIVE
def objectivefunction_rule(model):
    return sum( sum( sum( sum( 
                    model.Z[i,j,k,t] * (model.p[i,j,k,t] + model.pen[i,j,t]) 
            for i in model.i) for j in model.j) for k in model.k) for t in model.t)
model.OBJ = Objective(rule = objectivefunction_rule, sense = minimize)

instance = model.create_instance()

#CONSTRAINTS
#Capacity
def upper_capacity_rule(model,i,t):
    return sum( sum( model.Z[i,j,k,t] for j in model.j) for k in model.k) <= model.UB[i,t]
model.ConstraintUpperCapacity = Constraint(model.i, model.t, rule = upper_capacity_rule)
            
def lower_capacity_rule(model,i,t):
    return sum( sum( model.Z[i,j,k,t] for j in model.j) for k in model.k) >= model.LB[i,t]
model.ConstraintLowerCapacity = Constraint(model.i, model.t, rule = lower_capacity_rule)
            
#Price Intervals
def lower_price_interval_rule(model,i,k):
    return model.bin[i,k] * model.LT[i,k] <= sum( sum( model.X[i,j,k,t] for t in model.t) for j in model.j)
model.ConstraintLowerPrice = Constraint(model.i, model.k, rule = lower_price_interval_rule)
            
def upper_price_interval_rule(model,i,k):
    return model.bin[i,k] * model.UT[i,k] >= sum( sum( model.X[i,j,k,t] for t in model.t) for j in model.j)
model.ConstraintUpperPrice = Constraint(model.i, model.k, rule = lower_price_interval_rule)
            
def binary_sum_rule(model,i):
    return sum( model.bin[i,k] for k in model.k) == 1
model.ConstraintBinarySum = Constraint(model.i, rule = binary_sum_rule)
            
#z-Variable
def z_min_bound_rule(model,i,k,t):
    return sum( model.Z[i,j,k,t] for j in model.j) <= model.bin[i,k] * model.UB[i,t]
model.Constraint_z_min_bound = Constraint(model.i, model.k, model.t, rule = z_min_bound_rule)
            
def z_max_bound_rule(model,i,k,t):
    return sum( model.Z[i,j,k,t] for j in model.j) >= model.bin[i,k] * model.LB[i,t]
model.Constraint_z_max_bound = Constraint(model.i, model.k, model.t, rule = z_max_bound_rule)
            
def z_min_function_rule(model,i,k,t):
    return sum( model.Z[i,j,k,t] for j in model.j) <= sum( model.X[i,j,k,t] for j in model.j) - model.LB[i,t] * (1 - model.bin[i,k])
model.Constraint_z_min_function = Constraint(model.i, model.k, model.t, rule = z_min_function_rule)
            
def z_max_function_rule(model,i,k,t):
    return sum( model.Z[i,j,k,t] for j in model.j) >= sum( model.X[i,j,k,t] for j in model.j) - model.UB[i,t] * (1 - model.bin[i,k])
model.Constraint_z_max_function = Constraint(model.i, model.k, model.t, rule = z_max_function_rule)

#Forecasted Volume
def forecast_volume_rule(model,j,t):
    return sum( sum( model.X[i,j,k,t] for i in model.i) for k in model.k) >= model.d[j,t]
model.Constraint_forecast_volume = Constraint(model.j, model.t, rule = forecast_volume_rule)